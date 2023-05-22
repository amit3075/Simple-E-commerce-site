"""print("hello word")
message = 'hello_s word dedff'


print(message[0])
print((len(message)))
new_message =message.replace("hello", "Amit")
print(new_message)
greeting ='hello'
name = 'Amit'
message1 = greeting + ',' + name + '.welcome!'
message2 ='{},{}.welcome'.format(greeting,name)
message3 =f'{greeting},{name}.Welcome!'
print(message1)
print(message2)
print(message3)
num = 3.12
print(7%7)
print(13%7)
print(22%7)
print(2 + 4 * 5 / 5)
x = 1

x = x+1
x =-2
x += 3
print(x)
print(abs(-9))
print(round(3.987654321,5))
num_1 =2
num_3 = 3
print(num_1 == 2)
print(num_1>=num_3)
print(num_1<=num_3) 
num1 ='1000'
num2 ='2032'
num1 = int(num1)
num2 = int(num2)
print(num1 + num2)
#list example
courses = [ 'History', 'Math','Pysics','CompSci']
courses2 =['BHM','Cs','Humantity']
print(courses[-1])
print(courses[0:3])
courses.append('art')
courses.insert(3, 'sci')
courses.extend(courses2)
courses.remove('Math')
poped =courses.pop()
print(poped)
print(courses[-1])
print(courses)
courses = ['math','sci','Neapli']
num = [1,2,34,4,5,5]
courses.sort(reverse=True)
num.sort(reverse=True)
print(courses)
print(num)
num1 =[242541,253522,33222,3232,5252345235,1,]
sorted_num = sorted(num1)
print(sorted_num)
courses = ['history','math','sci','art']
print(courses.index('art'))
print('math' in courses)
for item in courses:
    print(item)
for index, courses in enumerate(courses):
    print(index,courses)
courses = ['math','sci','Cs']
courses_str = ','.join(courses)
new_list = courses_str.split(',')
print(courses_str)
list_1 = ['hello','world','to','amit']
list_2 = list_1
print(list_1)
print(list_2)
list_1[0] ='art'
print(list_1)
tuple_1 =('history','math','sci','cs')
tuple_2 = tuple_1
print(tuple_2)
print(tuple_1)
cs_courses = {'history','math','sci','cs'}
print('math' in cs_courses)
cs_courses1 ={ 'art','cs','music'}
print(cs_courses.intersection(cs_courses1))
print(cs_courses.difference(cs_courses1))
print(cs_courses.union(cs_courses1))
empty_list =[]
empty_list = list()
empty_tuple =()
empty_tuple = tuple()
empty_set = {}#this is not right:its a dict
empty_set =set()
print(empty_tuple)
empty_list.insert(0,'amit')
empty_list.insert(1,'hello')
print(empty_list)
#dict
student = {'name':'amit','age':24,'courses':['math','sci']}
student['phone'] ='3434434343'
student['name']='omit'
student.update({'name': 'aamshu','age':34,'courses':['neapli','cs'],'phone':'21345'})
print(student['name'])
print(student.get('phone','not found'))
age = student.pop('age')
print(len(student))
print(student.items())
print(student.values())
for key,value in student.items():
      print(key,value)
print(student)
#conditionals and boolean
language = input('hello ebter the')
if language == 'python':
    print('language was python')
elif language =='java':
    print('the lanaguage is java')
else:
    print('no match')
user ='Admin'
logged_in = False
if user =='Admin' or logged_in:
    print('admin Page')
else:
    print('welcome')
condition = None

if condition:
    print('evalutaed to ture')
else:
    print('evaluated to flase')
nums = [1,2,3,4]
for num in nums:
    for letter in 'abcd':
        print(num,letter)
    if num ==3:
        print("found!")
        continue
    print(num)

import my_module1 as mm
courses =['History','Math','Compsic']

index = mm.find_index(courses,'Math')
print(index)

        
from my_module1 import find_index as fi,test
courses =['History','Math','Compsic']

index = fi(courses,'Math')
print(index)
print(test)
from my_module1 import *
courses =['History','Math','Compsic']

index = find_index(courses,'Math')
print(index)
print(test)
import random
courses =['History','Math','Compsic']
random_courses = random.choices(courses)
print(random_courses)
import datetime
import calendar
today =datetime.date.today()
print(today)
print(calendar.isleap(2020))
import os
print(os.getcwd())
print(os.__file__)
print("hello")
print("first Module'")
def main():
    print ("first module's Name")

if __name__ =='__main__':
    main()

def functionName():
    print('HEllo world')

functionName()

#work with excel file inside the project
from openpyxl import Workbook, load_workbook

wb = load_workbook('Grade.xlsx')

WS = wb.active
print(WS)
print(WS['A5'].value)
#WS['A2'].value = "Test"

wb.create_sheet("Test")
print(wb.sheetnames)
#wb.save('Grade.xlsx')
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('Grade.xlsx')
WS = wb.active

for row in range(1,3):
    for col in range(1,3):
       # char = cahr(65 + Col)
       char = get_column_letter(col)
       print(WS[char + str(row)].value)


wb.save('Grade.xlsx')

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Grade.xlsx')
WS = wb.active

#WS.merge_cells('A1:D1')
WS.move_range("C1:D11",rows =2,cols=2)
wb.save('Grade.xlsx')

from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter


data ={
    "joe":{
        "math":65,
        "science":78,
        "english":88,
        "gym":78
    },
    "Bill":{
         "math":65,
         "science":78,
         "english":88,
         "gym":78

    },
    "Hari":{
         "math":65,
         "science":78,
         "english":88,
         "gym":78
    },
    "Amit":{ 
        "math":65,
        "science":78,
        "english":88,
        "gym":78
    }

}

wb = Workbook()
WS = wb.active
wb.title  = "Grades"

headings = ['Name'] + list( data['joe'].key())
ws.append(headings)

for person in data:
    grades =list(data[person].value())
    WS.append([person] + grades)

wb.save('Grade.xlsx')
#python OOP
class Dog:
    attr1 ="mamal"
    attr2 ="dog"
    def fun(self):
        print("i am ",self.attr1)
        print("i am a ",self.attr2)

Rodger = Dog()


#acessing class atribute
print(Rodger.attr1)
Rodger.fun()
# sample class with init method
class Person:
    #init method or constructor
    def __init__(self,name):
        self.name = name
    #sample Method
    def say_hello(self):
        print('hello, my name is ',self.name)

p = Person('AMit')
p.say_hello()
#class example
class Dog:
    #class var
    animal ='dog'

    #The init method or constructor
    def __init__(self,breed,color):
        #instance var
        self.breed =breed
        self.color =color
#obj of Dog class
Rodger = Dog("pug","Brown")
japniz = Dog("spazizes","whit")

print('Rodger details:')
print('Rodger is a ',Rodger.animal)
print('Breed:',Rodger.breed)
print("color",Rodger.color)
#class var can be accessed using class
#name also
print('\n Acessing class var using class name')
print(Dog.animal)
#instance var inside methods
#class for dog
class Dog:
    #class var
    animal ='dog'
    #init method or constructor
    def __init__(self,breed):
        #instance var
        self.breed =breed
    #adding an instance var
    def setColor(self,color):
        self.color =color

    #Retrives instance var
    def getcolor(self):
        return self.color
#Drive code
Rodger = Dog("pug")
Rodger.setColor(('Red'))
print(Rodger.getcolor())
class ClassName:
    x=0
    y=""

    def __init__(self, anyNumber, anystring):
        self.x = anyNumber
        self.y = anystring
myObject = ClassName(12345,"Hello")

print(myObject.__str__())
print(myObject.__repr__())
print(myObject)
class Myclass:
    x=0
    y=""

    def __init__(self,anyNumber,anyString):
        self.x= anyNumber
        self.y=anyString
    def __str__(self):
        return 'Myclass(x=' + str(self.x) +',y='+ self.y +')'

myobject = Myclass(12344, "ring")

print(myobject.__str__())
print(myobject)
print(str(myobject))
print(myobject.__repr__())
class Myclass:
    x=0
    y=""

    def __init__(self,anyNumber,anyString):
        self.x =anyNumber
        self.y =anyString
    def __repr__(self):
        return 'Myclass(x=' + str(self.x) +',y='+self.y+')'
myobject = Myclass(1234, "Stringfdsd")

print(myobject.__str__())
print(myobject)
print(str(myobject))
print(myobject.__repr__())
#Errror code
class InputNumber:
    def __init__(self):
        pass
    def getInpu(self,Number):
        a = int(input("enter the Number"))
        for i in a:
             print('helo')
a1 = InputNumber()#Error code
sport = input("enter a sport:")
p1_Score = int(input("enter the player 1 score:"))
p2_Score = int(input("Enter the score of second player:"))
#
if sport.lower() == "basketball":
    if p1_Score ==p2_Score:
         print("the Game is Draw")
    elif p1_Score < p2_Score:
        print("Player second win the game")
    elif p2_Score < p1_Score:
        print("player one win the game")
#loop for password
password =""
while password != "secret":
    password = input("enter the passwoed")
    if password == "ecret":
        print("correct password")
    else:
        print("password is incorect")
        

#creat the list 
list = ["DBMS","DSL",10]
print("list befor append()method")
print(list)
#adding single element in list
list.append(20)
#print the list
print("list after the append() method")
print(list)
list2 =["big data ",30,"ML",30]
list.extend(list2)
print("list aftert extend() method")
print(list)
for a in list:
    print(a)
#list
a ="ambition"
for character in a:
    print(character)
a ="ambition"
empty_list =list()
for character in a:
    empty_list.append(character)
#error look later
#comperhensive in list
a ="ambition"
empty_list =[character for character in a if character.lower()=='i']
print(character)
#empty dict
a = dict()
b = dict(red=1,blue=2,green=3)
print(b)
class SampleClass:
    #default constructor which is initialize the val
    def __init__(self):
        #key are initalized with
        #thir respective val
        self.exmp1 ='hello'
        self.exmp2 ="this"
        self.exmp3 ="is"
        self.exmp4 ="Bscist"
     #printDict method which print some sample text
    def printDict(self):
        print("dict const from obj field of the class sampleclass")
#crearting an obj to represent the class
sampleobj = SampleClass()
#calling printit method
sampleobj.printDict()
#calling obj __dict__ on sampleclass obj and print it
print(sampleobj.__dict__)
#ask for the sir
#for count,key in enumerate(a):
#    print(f"{count}:{key}")
a = [1,2,3,4,5,6,7,5,4,5,6]
set(a)
print(list(set(a)))
class Employe:
    def __init__(self,first:str,last:str,pay:int,batch=None):
        self.first = first#class instance
        self.last = last
        self.pay =pay
        self.email = f"{first}{last}@company.com"
        self.batch = batch
    def __repr__(self):
        return f"{self.first} has the {self.last}"
    def fullName(self):#instance method
        return f"{self.first} {self.last}"
emp1 = Employe("ram", "st",20000)
print(emp1)
print(dir(emp1))
print(emp1.first)
print(emp1.email)
print(emp1.fullName())
from my_Module1 import InputNumber as mm
courses =['History','Math','Compsic']

index = mm.find_index(courses,'Math')
print(index)
password =10
var1 = mm.getInputs()
if (password == var1):
    print("hello")
else:
    print('how are you') 
    #file handling

f= open("Hello.txt",'r')    
print(f)#fileobj
print(f.read())
f.close()
with open('Hello.txt','w') as file:
    print(file.read())
    file.readline()
    file.close()
with open('Hello.txt','r') as file:
    file.write('dfghjkldfghnjjjjjjjjjj')
    print(file.read())
    file.close()
with open('Hello.txt','r+') as f:
    print(f.read())
    print(help(f.seek))
    print(f.seek(6,0))
    print(f.read())
import io
std =io.StringIO()
print(std.write("hello world/n"))
print(std.getvalue())
#wrapping a file interface around a string
std = io.StringIO("hello\nworld\n")
print(std.readline())
 """




